import xlrd
import os.path
from os import path
from tkinter import *
from tkinter import filedialog
from openpyxl import *
from datetime import *
import datetime
import pathlib
import tkinter.font as tkFont

class Win:

  def __init__(self, root):
    
    self.root = root
    self.root.geometry("1020x770")
    self.background_color = "#107896"
    self.text_color = "#DFDFDF"
    self.text_color2 = "#a9ebba"
    self.text_color3 = "#c7941e"
    self.fontStyle = tkFont.Font(family="Lucida Grande", size=25)
    self.fontStyleMain = tkFont.Font(family="Lucida Grande", size=40)
    self.background_color2 = "#239BA9"
    self.root["bg"] = self.background_color
    #self.root.bind('<Return>', self.excel_click)
    self.main_title = Label(root, text = 'Coronavirus Quarantine Tracker', bg=self.background_color, fg=self.text_color, font=self.fontStyleMain)
    self.main_title.grid(row=0, column=0, padx=40, pady=45, columnspan=9)
    self.excel_button = Button(root, text="Choose Excel File", command=self.excel_click, bg=self.background_color2, fg=self.text_color2, font=self.fontStyle)
    self.excel_button.grid(row=1, column=0, padx=40, pady=15, columnspan=9)
    self.year_label = Label(root, text="Year :", bg=self.background_color, fg=self.text_color, font=self.fontStyle)
    self.year_entry = Entry(root, width=20, borderwidth=2, bg=self.background_color2, fg=self.text_color, font=self.fontStyle)
    self.month_label = Label(root, text="Month :", bg=self.background_color, fg=self.text_color, font=self.fontStyle)
    self.month_entry = Entry(root, width=20, borderwidth=2, bg=self.background_color2, fg=self.text_color, font=self.fontStyle)
    self.day_label = Label(root, text="Day :", bg=self.background_color, fg=self.text_color, font=self.fontStyle)
    self.day_entry = Entry(root, width=20, borderwidth=2, bg=self.background_color2, fg=self.text_color, font=self.fontStyle)
    self.submit_btn = Button(root, text="Submit", command=self.data_click, bg=self.background_color2, fg=self.text_color2, font=self.fontStyle)
    self.no_file = Label(root, text="No file found. Please try again.", bg=self.background_color, fg=self.text_color, font=self.fontStyle)

		
 
  def excel_click(self):
    global filename 
    filename = filedialog.askopenfilename(initialdir = "/Desktop", title="Select the excel file", filetypes=[("Excel files", ".xlsx .xls")])
    if os.path.exists(filename): 
      #self.root.bind('<Return>', self.data_click(self))
      self.success_label = Label(self.root, text="File chosen successfully", bg=self.background_color, fg=self.text_color, font=self.fontStyle)
      self.success_label.grid(row=2, column=0, padx=10, pady=15, columnspan=9)
      self.date_ask = Label(self.root, text="Enter today's date in this format: yyyy/mm/dd and then click Submit", bg=self.background_color, fg=self.text_color, font=self.fontStyle)
      self.date_ask.grid(row=3, column=0, columnspan=9, padx=20, pady=20)
      self.year_label.grid(row=4, column=0, padx=0, pady=15)
      self.year_entry.grid(row=4, column=1, padx=0, pady=15)
      self.month_label.grid(row=5, column=0, padx=0, pady=15)
      self.month_entry.grid(row=5, column=1, padx=0, pady=15)
      self.day_label.grid(row=6, column=0, padx=0, pady=15)
      self.day_entry.grid(row=6, column=1, padx=0, pady=15)
      self.submit_btn.grid(row=7, column=0, padx=10, pady=15, columnspan=9)
    else: 
      self.no_file.grid(row=2, column=0, padx=10, pady=15)
   
  def data_click(self):
    if (int(self.year_entry.get()) >= 2020 and int(self.year_entry.get()) <= 2022) and (int(self.month_entry.get()) >= 0 and int(self.month_entry.get()) <= 12) and (int(self.day_entry.get()) >= 0 and int(self.day_entry.get()) <= 31):
      global year, month, day
      year = int(self.year_entry.get())
      month = int(self.month_entry.get())
      day = int(self.day_entry.get())
      self.year_entry.delete(0, END)
      self.month_entry.delete(0, END)
      self.day_entry.delete(0, END)
      wb = load_workbook(filename=filename)
      sheet = wb.worksheets[0]
      row_count = sheet.max_row
      ending_count = 0
      quarantine_count = 0
      new_filename = "{}.{}.{} Coronavirus quarantine report.xlsx".format(year, month, day)
      workbook = Workbook()
      new_sheet = workbook.active
      row_index = 2
      letters = "ZABCDEFGHIJKLMNOP"
    
      for i in range(5, row_count):
        if type(sheet.cell(row=i, column=15).value) is datetime.datetime:
          start_date = sheet.cell(row=i, column=15).value
          end_date = sheet.cell(row=i, column=15).value

          if int(end_date.strftime("%Y")) == year and (int(end_date.strftime("%m")) > month or (int(end_date.strftime("%d")) > day) and int(end_date.strftime("%m")) == month):
            quarantine_count += 1
            row_index += 1
            for j in range(1, 16):
              new_sheet['{}{}'.format( letters[j] , row_index)] =  sheet.cell(row=i, column=j).value
              
        
        elif type(sheet.cell(row=i, column=15).value) == str:
          start_date = sheet.cell(row=i, column=15).value.split('.')
          end_date = sheet.cell(row=i, column=15).value.split('.')
          try:
            if int(end_date[0]) == year and (int(end_date[1]) > month or (int(end_date[2]) > day and int(end_date[1]) == month)):
              quarantine_count += 1
              row_index += 1
              for j in range(1, 16):
                new_sheet['{}{}'.format( letters[j] , row_index)] =  sheet.cell(row=i, column=j).value
          except:
            pass

      new_sheet['A2'] = "The {} people that need to be quarantined on {}.{}.{}".format(quarantine_count, year, month, day)
      row_index += 1
      row_placeholder = row_index
      
      for i in range(5, row_count):
        if type(sheet.cell(row=i, column=15).value) is datetime.datetime:
          start_date = sheet.cell(row=i, column=15).value
          end_date = sheet.cell(row=i, column=15).value

          if int(end_date.strftime("%Y")) == year and int(end_date.strftime("%m")) == month and int(end_date.strftime("%d")) == day:
            ending_count += 1
            row_index += 1
            for j in range(1, 16):
              new_sheet['{}{}'.format( letters[j] , row_index)] =  sheet.cell(row=i, column=j).value
        
        elif type(sheet.cell(row=i, column=15).value) == str:
          start_date = sheet.cell(row=i, column=15).value.split('.')
          end_date = sheet.cell(row=i, column=15).value.split('.')
          try:
            if int(end_date[0]) == year and int(end_date[1]) == month and int(end_date[2]) == day:
              ending_count += 1
              row_index += 1
              for j in range(1, 16):
                new_sheet['{}{}'.format( letters[j] , row_index)] =  sheet.cell(row=i, column=j).value
          except:
            pass
      
      new_sheet['A{}'.format(row_placeholder)] = "The {} people that are leaving quarantine on {}.{}.{}".format(ending_count, year, month, day)
      workbook.save(filename=new_filename)
      path = pathlib.Path(__file__).parent.absolute()
      os.system('start "excel" "{}//{}"'.format(path, new_filename))
		

def main():
	root = Tk()
	root.title('Khan Uul Corona Quarantine Tracker')
	app = Win(root)
	root.mainloop()








if __name__ == "__main__":
	main()















		
 
 
 


