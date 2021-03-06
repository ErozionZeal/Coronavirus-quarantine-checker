import xlrd
import os.path
from os import path
from tkinter import *
from tkinter import filedialog
from openpyxl import *
from datetime import *
import datetime
import os
import pathlib

class Win:

	def __init__(self, root):


		self.root = root
		self.root.geometry("800x1200")
		self.root["bg"] = "coral"
		#self.main = tk.Frame(self.root)
		#self.main.pack(fill=tk.BOTH, expand=True)
		self.main_title = Label(root, text = 'Coronavirus Quarantine Tracker')
		self.main_title.place(x = 20, y = 30, width=720, height=175)
		self.excel_button = Button(root, text="Choose Excel File", command=self.excel_click)
		self.excel_button.place(x = 20, y = 215, width=360, height=125)
		self.year_label = Label(root, text="Year :")
		self.year_entry = Entry(root, width=20, borderwidth=5)
		self.month_label = Label(root, text="Month :")
		self.month_entry = Entry(root, width=20, borderwidth=5)
		self.day_label = Label(root, text="Day :")
		self.day_entry = Entry(root, width=20, borderwidth=5)
		self.submit_btn = Button(root, text="Submit", command=self.data_click)
		self.no_file = Label(root, text="No file found. Please try again")

		
 
	def excel_click(self):
		global filename 
		filename = filedialog.askopenfilename(initialdir = "/Desktop", title="Select the excel file", filetypes=[("Excel files", ".xlsx .xls")])
		if os.path.exists(filename): 
			self.success_label = Label(self.root, text="File chosen successfully")
			self.success_label.place(x = 320, y = 350, width=220, height=75)
			self.date_ask = Label(self.root, text="Enter today's date in this format: yyyy/mm/dd and then click Submit")
			self.date_ask.place(x = 320, y = 430, width=220, height=75)
			self.year_label.place(x = 20, y = 500, width=120, height=50)
			self.year_entry.place(x = 150, y = 500, width=100, height=50)
			self.month_label.place(x = 260, y = 500, width=120, height=50)
			self.month_entry.place(x = 390, y = 500, width=80, height=50)
			self.day_label.place(x = 480, y = 500, width=120, height=50)
			self.day_entry.place(x = 610, y = 500, width=80, height=50)
			self.submit_btn.place(x = 20, y = 750, width=360, height=125)
			#place(x = 320, y = 570, width=220, height=75)
		else: 
			self.no_file.place(x = 320, y = 350, width=220, height=75)


	def data_click(self):
		if (int(self.year_entry.get()) >= 2020 and int(self.year_entry.get()) <= 2022) and (int(self.month_entry.get()) >= 0 and int(self.month_entry.get()) <= 12) and (int(self.day_entry.get()) >= 0 and int(self.day_entry.get()) <= 31):
			global year, month, day
			year = int(self.year_entry.get())
			month = int(self.month_entry.get())
			day = int(self.day_entry.get())
			self.year_entry.delete(0, END)
			self.month_entry.delete(0, END)
			self.day_entry.delete(0, END)
			self.new_window(Win2)
			
			
   
	def new_window(self, _class):
  		
		try:
      
			if self.new.state() == "normal":
				self.new.focus()
		except:
			self.new = Toplevel(self.root)
			_class(self.new)
 
class Win2:
	def __init__(self, root):
		
		wb = load_workbook(filename=filename)
		sheet = wb.worksheets[0]
		row_count = sheet.max_row
		ending_count = 0
		quarantine_count = 0
		new_filename = "{}.{}.{} Coronavirus quarantine report.xlsx".format(year, month, day)
		workbook = Workbook()
		new_sheet = workbook.active
		row_index = 2
		
  
		for i in range(5, row_count):
			if type(sheet.cell(row=i, column=15).value) is datetime.datetime:
				start_date = date.fromisoformat(sheet.cell(row=i, column=15).value)
				end_date = date.fromisoformat(sheet.cell(row=i, column=15).value)

				if int(end_date.strftime("%Y")) == year and (int(end_date.strftime("%m")) > month or int(end_date.strftime("%d")) > day):
					quarantine_count += 1
					row_index += 1
					for j in range(1,16):
						new_sheet.cell(row=row_index, column=j) =  sheet.cell(row=i, column=j).value
						
			
			elif type(sheet.cell(row=i, column=15).value) == str:
				start_date = sheet.cell(row=i, column=15).value.split('.')
				end_date = sheet.cell(row=i, column=15).value.split('.')
				try:
					if int(end_date[0]) == year and (int(end_date[1]) > month or int(end_date[2]) > day):
						quarantine_count += 1
						row_index += 1
						for j in range(1,16):
							new_sheet.cell(row=row_index, column=j) =  sheet.cell(row=i, column=j).value
				except:
					pass

		new_sheet.cell(row = 2, column=j) = "The {} people that need to be quarantined on {}.{}.{}".format(quarantine_count, year, month, day)
		row_placeholder = row_index + 1
     
		for i in range(5, row_count):
			if type(sheet.cell(row=i, column=15).value) is datetime.datetime:
				start_date = date.fromisoformat(sheet.cell(row=i, column=15).value)
				end_date = date.fromisoformat(sheet.cell(row=i, column=15).value)

				if int(end_date.strftime("%Y")) == year and int(end_date.strftime("%m")) and int(end_date.strftime("%d")) == day:
					ending_count += 1
					row_index += 1
					for j in range(1,16):
						new_sheet.cell(row=row_index, column=j) =  sheet.cell(row=i, column=j).value
			
			elif type(sheet.cell(row=i, column=15).value) == str:
				start_date = sheet.cell(row=i, column=15).value.split('.')
				end_date = sheet.cell(row=i, column=15).value.split('.')
				try:
					if int(end_date.strftime("%Y")) == year and int(end_date.strftime("%m")) and int(end_date.strftime("%d")) == day:
						ending_count += 1
						row_index += 1
						for j in range(1,16):
							new_sheet.cell(row=row_index, column=j) =  sheet.cell(row=i, column=j).value
				except:
					pass
		 
		new_sheet.cell(row = row_placeholder, column=j) = "The {} people that are leaving quarantine on {}.{}.{}".format(ending_count, year, month, day)
		workbook.save(filename=new_filename)
		path = pathlib.Path(__file__).parent.absolute()
		os.system('start "excel" "{}{}"'.format(path, new_filename))
		

def main():
	root = Tk()
	root.title('Khan Uul Corona Quarantine Tracker')
	app = Win(root)
	root.mainloop()








if __name__ == "__main__":
	main()















		
 
 
 


